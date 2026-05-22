from __future__ import annotations

import argparse
import csv
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_ROOT = ROOT / "tests" / "sample_artifacts"


def _ensure_dirs() -> dict[str, Path]:
    paths = {
        "valid_excel": OUTPUT_ROOT / "valid" / "excel",
        "valid_csv": OUTPUT_ROOT / "valid" / "csv",
        "valid_parquet": OUTPUT_ROOT / "valid" / "parquet",
        "invalid_excel": OUTPUT_ROOT / "invalid" / "excel",
        "invalid_csv": OUTPUT_ROOT / "invalid" / "csv",
        "invalid_parquet": OUTPUT_ROOT / "invalid" / "parquet",
    }
    for p in paths.values():
        p.mkdir(parents=True, exist_ok=True)
    return paths


def _load_valid_transaction_base(valid_csv_dir: Path) -> pd.DataFrame:
    return pd.read_csv(valid_csv_dir / "valid_transactions_001.csv")


def _load_valid_customer_base_frames(valid_excel_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    base_file = valid_excel_dir / "valid_customers_001.xlsx"
    au_df = pd.read_excel(base_file, sheet_name="Customers_AU", engine="openpyxl")
    us_df = pd.read_excel(base_file, sheet_name="Customers_US", engine="openpyxl")
    return au_df, us_df


def _set_excel_column_text_format(file_path: Path, sheet_name: str, column_name: str) -> None:
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        wb.close()
        return

    col_idx = headers.index(column_name) + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None and not isinstance(cell.value, str):
            cell.value = str(cell.value)
        cell.number_format = "@"

    wb.save(file_path)
    wb.close()


def _build_customers_rows(file_seq: int, start_id: int, count: int, region: str) -> list[dict]:
    rows: list[dict] = []
    base_date = datetime(2024, 1, 1) + timedelta(days=file_seq * 7)

    for i in range(count):
        idx = start_id + i
        signup_date = base_date + timedelta(days=i)
        rows.append(
            {
                "CustomerId": f"CUST{idx:05d}",
                "CustomerName": f"Customer {idx}",
                "SignupDate": signup_date.date(),
                "CreditLimit": round(1500 + (i * 35.75) + (file_seq * 50), 2),
                "IsActive": "Y" if i % 5 != 0 else "N",
                "RegionCode": region,
                "SourceSystem": "CRM",
            }
        )
    return rows


def _apply_signup_date_number_format(writer: pd.ExcelWriter, sheet_name: str, format_code: str) -> None:
    ws = writer.book[sheet_name]
    headers = [cell.value for cell in ws[1]]
    if "SignupDate" not in headers:
        return

    signup_col_idx = headers.index("SignupDate") + 1
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=signup_col_idx).number_format = format_code


def _apply_column_date_number_format(
    writer: pd.ExcelWriter,
    sheet_name: str,
    column_name: str,
    format_code: str,
) -> None:
    ws = writer.book[sheet_name]
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        return

    col_idx = headers.index(column_name) + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        # Only style true date/datetime cells; leave string error cases unchanged.
        if hasattr(cell.value, "year") and hasattr(cell.value, "month") and hasattr(cell.value, "day"):
            cell.number_format = format_code


def _apply_column_text_number_format(
    writer: pd.ExcelWriter,
    sheet_name: str,
    column_name: str,
) -> None:
    ws = writer.book[sheet_name]
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        return

    col_idx = headers.index(column_name) + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None and not isinstance(cell.value, str):
            cell.value = str(cell.value)
        cell.number_format = "@"


def generate_valid_excel_files(valid_excel_dir: Path) -> None:
    for n in (1, 2, 3):
        file_path = valid_excel_dir / f"valid_customers_{n:03d}.xlsx"

        au_rows = _build_customers_rows(file_seq=n, start_id=1000 + (n * 100), count=8, region="AU")
        us_rows = _build_customers_rows(file_seq=n, start_id=2000 + (n * 100), count=8, region="US")

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            pd.DataFrame(au_rows).to_excel(writer, sheet_name="Customers_AU", index=False)
            pd.DataFrame(us_rows).to_excel(writer, sheet_name="Customers_US", index=False)

            # Explicit display formats per worksheet while preserving true Excel date values.
            _apply_signup_date_number_format(writer, "Customers_AU", "d/mm/yyyy;@")
            _apply_signup_date_number_format(writer, "Customers_US", "m/d/yyyy;@")


def generate_valid_csv_files(valid_csv_dir: Path) -> None:
    file_1 = valid_csv_dir / "valid_transactions_001.csv"
    file_2 = valid_csv_dir / "valid_transactions_002.csv"
    file_large = valid_csv_dir / "valid_transactions_large.csv"

    def _rows(offset: int, count: int) -> list[dict]:
        rows: list[dict] = []
        for i in range(count):
            idx = offset + i
            rows.append(
                {
                    "TransactionId": f"TXN{idx:06d}",
                    "CustomerId": f"CUST{1000 + (idx % 50):05d}",
                    "TransactionDate": (datetime(2025, 1, 1) + timedelta(days=i % 90)).date().isoformat(),
                    "Amount": round(20 + (i * 1.87), 2),
                    "Currency": "AUD" if i % 2 == 0 else "USD",
                    "Status": "COMPLETE",
                }
            )
        return rows

    pd.DataFrame(_rows(offset=1, count=25)).to_csv(file_1, index=False)
    pd.DataFrame(_rows(offset=1000, count=25)).to_csv(file_2, index=False)

    # Large CSV test fixture for chunking + boundary behavior:
    # - 1,000,000 data rows
    # - 20 source columns
    # - two leading non-data rows before header (header_skip_rows=2)
    large_columns = [
        "TransactionId",
        "CustomerId",
        "TransactionDate",
        "Amount",
        "Currency",
        "Status",
        "Quantity",
        "DiscountRate",
        "FeeAmount",
        "TaxAmount",
        "NetAmount",
        "Channel",
        "Region",
        "SourceSystem",
        "BatchId",
        "EventTimestamp",
        "IsPriority",
        "ReferenceCode",
        "LedgerCode",
        "CommentText",
    ]

    with file_large.open("w", encoding="utf-8", newline="") as fh:
        fh.write("# preamble row 1 for skip-rows testing\n")
        fh.write("# preamble row 2 for skip-rows testing\n")
        writer = csv.DictWriter(fh, fieldnames=large_columns)
        writer.writeheader()

        for i in range(1_000_000):
            idx = 100000 + i
            txn_date = (datetime(2025, 1, 1) + timedelta(days=i % 365)).date().isoformat()
            amount = round(20 + (i * 0.0137), 2)
            discount_rate = round((i % 15) / 100, 4)
            fee_amount = round(amount * 0.0125, 2)
            tax_amount = round(amount * 0.10, 2)
            net_amount = round(amount - fee_amount + tax_amount, 2)

            writer.writerow(
                {
                    "TransactionId": f"TXN{idx:07d}",
                    "CustomerId": f"CUST{1000 + (idx % 500):05d}",
                    "TransactionDate": txn_date,
                    "Amount": amount,
                    "Currency": "AUD" if i % 2 == 0 else "USD",
                    "Status": "COMPLETE" if i % 10 != 0 else "PENDING",
                    "Quantity": (i % 50) + 1,
                    "DiscountRate": discount_rate,
                    "FeeAmount": fee_amount,
                    "TaxAmount": tax_amount,
                    "NetAmount": net_amount,
                    "Channel": "ONLINE" if i % 3 == 0 else "STORE",
                    "Region": "AU" if i % 2 == 0 else "US",
                    "SourceSystem": "ERP",
                    "BatchId": f"BATCH{(i // 10000) + 1:05d}",
                    "EventTimestamp": f"{txn_date}T{(i % 24):02d}:{(i % 60):02d}:{((i * 3) % 60):02d}",
                    "IsPriority": "Y" if i % 20 == 0 else "N",
                    "ReferenceCode": f"REF{idx:09d}",
                    "LedgerCode": f"LED{(i % 200):03d}",
                    "CommentText": f"Large transaction row {idx}",
                }
            )


def generate_invalid_csv_files(invalid_csv_dir: Path, valid_csv_dir: Path) -> None:
    base = _load_valid_transaction_base(valid_csv_dir)

    mixed_types = base.head(3).copy()
    mixed_types["Amount"] = mixed_types["Amount"].astype("object")
    mixed_types["TransactionDate"] = mixed_types["TransactionDate"].astype("object")
    mixed_types["Status"] = mixed_types["Status"].astype("object")
    mixed_types.loc[mixed_types.index[0], "Amount"] = "ABC"
    mixed_types.loc[mixed_types.index[1], "TransactionDate"] = "not-a-date"
    mixed_types.loc[mixed_types.index[2], "Amount"] = "2025-03-01"
    mixed_types.loc[mixed_types.index[2], "Status"] = "123"
    mixed_types.to_csv(invalid_csv_dir / "invalid_mixed_types.csv", index=False)

    not_null_and_missing = base.head(3).copy()
    if "TransactionId" in not_null_and_missing.columns:
        not_null_and_missing = not_null_and_missing.drop(columns=["TransactionId"])
    not_null_and_missing["CustomerId"] = not_null_and_missing["CustomerId"].astype("object")
    not_null_and_missing["TransactionDate"] = not_null_and_missing["TransactionDate"].astype("object")
    not_null_and_missing["Amount"] = not_null_and_missing["Amount"].astype("object")
    not_null_and_missing.loc[not_null_and_missing.index[0], "CustomerId"] = ""
    not_null_and_missing.loc[not_null_and_missing.index[1], "TransactionDate"] = ""
    not_null_and_missing.loc[not_null_and_missing.index[2], "Amount"] = ""
    not_null_and_missing.to_csv(invalid_csv_dir / "invalid_not_null_and_missing_columns.csv", index=False)


def generate_invalid_excel_files(invalid_excel_dir: Path, valid_excel_dir: Path) -> None:
    base_au, base_us = _load_valid_customer_base_frames(valid_excel_dir)

    # 1) multiple datasets in same sheet + mixed type issues in numeric/date fields
    file_1 = invalid_excel_dir / "invalid_customers_multiple_datasets.xlsx"
    repeated_header = pd.DataFrame([dict(zip(base_au.columns, base_au.columns))])
    first_chunk = base_au.head(4).copy()
    second_chunk = base_au.head(2).copy()
    second_chunk["SignupDate"] = second_chunk["SignupDate"].astype("object")
    second_chunk["CreditLimit"] = second_chunk["CreditLimit"].astype("object")
    second_chunk.loc[second_chunk.index[0], "SignupDate"] = "31/13/2025"
    second_chunk.loc[second_chunk.index[1], "CreditLimit"] = "ABC"
    combined = pd.concat([first_chunk, repeated_header, second_chunk], ignore_index=True)
    with pd.ExcelWriter(file_1, engine="openpyxl") as writer:
        combined.to_excel(writer, sheet_name="Customers_AU", index=False)
        base_us.head(6).to_excel(writer, sheet_name="Customers_US", index=False)

    # 2) workbook intentionally missing expected tabs like Customers_AU / Customers_US
    file_2 = invalid_excel_dir / "invalid_missing_tabs.xlsx"
    with pd.ExcelWriter(file_2, engine="openpyxl") as writer:
        base_au.head(6).to_excel(writer, sheet_name="OnlySheet", index=False)

    # 3) additional unknown columns + likely truncation values
    file_3 = invalid_excel_dir / "invalid_additional_unknown_columns.xlsx"
    au_unknown = base_au.head(4).copy()
    au_unknown["SignupDate"] = au_unknown["SignupDate"].astype("object")
    au_unknown["CreditLimit"] = au_unknown["CreditLimit"].astype("object")
    au_unknown.loc[au_unknown.index[0], "CustomerName"] = "X" * 400
    au_unknown.loc[au_unknown.index[1], "SignupDate"] = "not-a-date"
    au_unknown.loc[au_unknown.index[2], "CreditLimit"] = "BAD_NUMBER"
    au_unknown["UnexpectedCode"] = ["UNMAPPED_001", "UNMAPPED_002", "UNMAPPED_003", "UNMAPPED_004"]
    au_unknown["ProcessName"] = ["does_not_exist"] * len(au_unknown)
    with pd.ExcelWriter(file_3, engine="openpyxl") as writer:
        au_unknown.to_excel(writer, sheet_name="Customers_AU", index=False)
        base_us.head(4).to_excel(writer, sheet_name="Customers_US", index=False)

    # 4) datetime stress workbook aligned to valid customer tabs
    file_4 = invalid_excel_dir / "invalid_datetime_stress.xlsx"
    au_stress = base_au.head(4).copy()
    us_stress = base_us.head(4).copy()
    au_stress["SignupDate"] = au_stress["SignupDate"].astype("object")
    us_stress["SignupDate"] = us_stress["SignupDate"].astype("object")
    au_stress.loc[au_stress.index[1], "SignupDate"] = "31/13/2025"
    au_stress.loc[au_stress.index[2], "SignupDate"] = "not-a-date"
    us_stress.loc[us_stress.index[1], "SignupDate"] = "13/31/2025"
    us_stress.loc[us_stress.index[2], "SignupDate"] = "2025-13-01"
    with pd.ExcelWriter(file_4, engine="openpyxl") as writer:
        au_stress.to_excel(writer, sheet_name="Customers_AU", index=False)
        us_stress.to_excel(writer, sheet_name="Customers_US", index=False)
        _apply_signup_date_number_format(writer, "Customers_AU", "d/mm/yyyy;@")
        _apply_signup_date_number_format(writer, "Customers_US", "m/d/yyyy;@")

    # 5) subtle invalid case: date-looking text values presented/stored as TEXT in Excel
    file_5 = invalid_excel_dir / "invalid_date_as_text.xlsx"
    au_text = base_au.head(4).copy()
    us_text = base_us.head(4).copy()
    au_text["SignupDate"] = au_text["SignupDate"].astype("object")
    us_text["SignupDate"] = us_text["SignupDate"].astype("object")
    au_text["SignupDate"] = ["31/01/2025", "01/02/2025", "15/03/2025", "2025-04-01"]
    us_text["SignupDate"] = ["01/31/2025", "02/01/2025", "03/15/2025", "2025-04-01"]
    with pd.ExcelWriter(file_5, engine="openpyxl") as writer:
        au_text.to_excel(writer, sheet_name="Customers_AU", index=False)
        us_text.to_excel(writer, sheet_name="Customers_US", index=False)
    _set_excel_column_text_format(file_5, "Customers_AU", "SignupDate")
    _set_excel_column_text_format(file_5, "Customers_US", "SignupDate")

    # 6) numeric precision/scale overflow scenarios (looks numeric but exceeds DECIMAL constraints)
    file_6 = invalid_excel_dir / "invalid_numeric_overflow.xlsx"
    overflow = base_au.head(3).copy()
    # precision overflow for DECIMAL(18,2): 19 total digits
    overflow.loc[overflow.index[0], "CreditLimit"] = 1234567890123456789
    # scale overflow for DECIMAL(*,2): 3 fractional digits
    overflow.loc[overflow.index[1], "CreditLimit"] = 123.456
    overflow.loc[overflow.index[2], "CreditLimit"] = 999.99
    with pd.ExcelWriter(file_6, engine="openpyxl") as writer:
        overflow.to_excel(writer, sheet_name="Customers_AU", index=False)
        base_us.head(3).to_excel(writer, sheet_name="Customers_US", index=False)


def _build_parquet_batch(offset: int, count: int, total_columns: int) -> pd.DataFrame:
    event_base = datetime(2025, 1, 1)
    rows: dict[str, list] = {
        "transaction_id": [f"TXN{offset + i:010d}" for i in range(count)],
        "customer_id": [f"CUST{100000 + ((offset + i) % 200000):06d}" for i in range(count)],
        "account_id": [f"ACC{50000 + ((offset + i) % 120000):06d}" for i in range(count)],
        "merchant_id": [f"MER{7000 + ((offset + i) % 9000):05d}" for i in range(count)],
        "transaction_date": [
            (event_base + timedelta(days=(offset + i) % 365)).date().isoformat()
            for i in range(count)
        ],
        "transaction_timestamp": [
            f"{(event_base + timedelta(days=(offset + i) % 365)).date().isoformat()}T{((offset + i) % 24):02d}:{((offset + i) % 60):02d}:{(((offset + i) * 3) % 60):02d}"
            for i in range(count)
        ],
        "amount": [round(100 + ((offset + i) % 10000) * 0.17, 2) for i in range(count)],
        "tax_amount": [round(10 + ((offset + i) % 1000) * 0.02, 2) for i in range(count)],
        "fee_amount": [round(1 + ((offset + i) % 500) * 0.01, 2) for i in range(count)],
        "net_amount": [round(90 + ((offset + i) % 11000) * 0.15, 2) for i in range(count)],
        "currency": ["AUD" if (offset + i) % 2 == 0 else "USD" for i in range(count)],
        "status": ["COMPLETE" if (offset + i) % 10 != 0 else "PENDING" for i in range(count)],
        "payment_method": ["CARD" if (offset + i) % 3 == 0 else "TRANSFER" for i in range(count)],
        "channel": ["ONLINE" if (offset + i) % 4 == 0 else "BRANCH" for i in range(count)],
        "country_code": ["AU" if (offset + i) % 2 == 0 else "US" for i in range(count)],
        "region_code": [f"R{((offset + i) % 12) + 1:02d}" for i in range(count)],
        "source_system": ["ERP" for _ in range(count)],
        "batch_reference": [f"BATCH{((offset + i) // 10000) + 1:06d}" for i in range(count)],
        "is_refund": [((offset + i) % 13 == 0) for i in range(count)],
        "risk_score": [round(((offset + i) % 1000) / 1000, 4) for i in range(count)],
        "device_id": [f"DEV{((offset + i) % 50000):05d}" for i in range(count)],
        "description": [f"Parquet transaction row {offset + i}" for i in range(count)],
    }

    current_columns = len(rows)
    if total_columns > current_columns:
        for n in range(current_columns + 1, total_columns + 1):
            rows[f"extra_col_{n:02d}"] = [f"X{n}_{(offset + i) % 1000:03d}" for i in range(count)]

    return pd.DataFrame(rows)


def generate_large_parquet_file(
    valid_parquet_dir: Path,
    *,
    target_size_gb: float,
    total_columns: int = 20,
    rows_per_batch: int = 50_000,
    filename: str | None = None,
) -> Path:
    if total_columns < 20:
        raise ValueError("total_columns must be at least 20")

    if target_size_gb <= 0:
        raise ValueError("target_size_gb must be greater than zero")

    resolved_filename = filename or f"valid_transactions_large_{str(target_size_gb).replace('.', '_')}gb.parquet"
    file_path = valid_parquet_dir / resolved_filename

    target_bytes = int(target_size_gb * 1024 * 1024 * 1024)
    rows_written = 0

    if file_path.exists():
        file_path.unlink()

    writer: pq.ParquetWriter | None = None
    try:
        while True:
            frame = _build_parquet_batch(rows_written, rows_per_batch, total_columns)
            table = pa.Table.from_pandas(frame, preserve_index=False)

            if writer is None:
                writer = pq.ParquetWriter(file_path, table.schema, compression="snappy")

            writer.write_table(table, row_group_size=min(rows_per_batch, 50_000))
            rows_written += len(frame)

            current_size = file_path.stat().st_size if file_path.exists() else 0
            if current_size >= target_bytes:
                break

            if rows_written % (rows_per_batch * 10) == 0:
                print(
                    f"Generating {file_path.name}: rows={rows_written:,}, "
                    f"size={current_size / (1024 * 1024 * 1024):.2f} GB"
                )
    finally:
        if writer is not None:
            writer.close()

    print(
        f"Generated {file_path.name}: rows={rows_written:,}, "
        f"size={file_path.stat().st_size / (1024 * 1024 * 1024):.2f} GB"
    )
    return file_path


def generate_valid_parquet_files(valid_parquet_dir: Path) -> None:
    valid_df = pd.DataFrame(
        {
            "transaction_id": ["PTX000001", "PTX000002", "PTX000003"],
            "customer_id": ["CUST01001", "CUST01002", "CUST01003"],
            "transaction_date": ["2025-01-15", "2025-01-16", "2025-01-17"],
            "amount": [125.50, 88.10, 230.00],
            "currency": ["AUD", "AUD", "USD"],
            "status": ["COMPLETE", "COMPLETE", "PENDING"],
            "source_system": ["ERP", "ERP", "ERP"],
        }
    )
    valid_df.to_parquet(valid_parquet_dir / "valid_transactions_parquet_001.parquet", index=False)


def generate_invalid_parquet_files(invalid_parquet_dir: Path) -> None:
    # Deliberately includes an over-length status value on the final row so
    # full-file chunked validation can detect an issue that appears after
    # the first chunk.
    invalid_df = pd.DataFrame(
        {
            "transaction_id": ["PTX900001", "PTX900002", "PTX900003"],
            "customer_id": ["CUST09901", "CUST09902", "CUST09903"],
            "transaction_date": ["2025-03-01", "2025-03-02", "2025-03-03"],
            "amount": [40.25, 50.75, 60.10],
            "currency": ["AUD", "USD", "AUD"],
            "status": ["OK", "GOOD", "STATUS_VALUE_EXCEEDS_LIMIT"],
            "source_system": ["ERP", "ERP", "ERP"],
        }
    )
    invalid_df.to_parquet(invalid_parquet_dir / "invalid_transactions_parquet_001.parquet", index=False)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate sample ingestion artifacts")
    parser.add_argument(
        "--large-parquet",
        action="store_true",
        help="Generate an additional large parquet artifact (opt-in)",
    )
    parser.add_argument(
        "--target-size-gb",
        type=float,
        default=2.0,
        help="Target parquet size in GB when --large-parquet is enabled",
    )
    parser.add_argument(
        "--parquet-columns",
        type=int,
        default=20,
        help="Total columns for generated large parquet artifact (minimum 20)",
    )
    parser.add_argument(
        "--parquet-rows-per-batch",
        type=int,
        default=50_000,
        help="Rows per batch while generating large parquet artifact",
    )
    args = parser.parse_args()

    paths = _ensure_dirs()
    legacy_datetime_csv = paths["invalid_csv"] / "invalid_datetime_stress.csv"
    if legacy_datetime_csv.exists():
        legacy_datetime_csv.unlink()
    generate_valid_excel_files(paths["valid_excel"])
    generate_valid_csv_files(paths["valid_csv"])
    generate_valid_parquet_files(paths["valid_parquet"])
    generate_invalid_csv_files(paths["invalid_csv"], paths["valid_csv"])
    generate_invalid_excel_files(paths["invalid_excel"], paths["valid_excel"])
    generate_invalid_parquet_files(paths["invalid_parquet"])

    if args.large_parquet:
        generate_large_parquet_file(
            paths["valid_parquet"],
            target_size_gb=args.target_size_gb,
            total_columns=max(20, args.parquet_columns),
            rows_per_batch=max(1_000, args.parquet_rows_per_batch),
        )

    print(f"Sample artifacts generated under: {OUTPUT_ROOT}")


if __name__ == "__main__":
    main()
